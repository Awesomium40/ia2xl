import openpyxl
from caastools.parsing import ia
from openpyxl.styles import Alignment, Border, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils import quote_sheetname
from openpyxl.worksheet.datavalidation import DataValidation

SIDE = Side(border_style='thin', color="000000")
BORDER = Border(SIDE, SIDE, SIDE, SIDE)
ALIGN = Alignment(horizontal="left", vertical="bottom", wrap_text=True, shrink_to_fit=False)


def _append_validation_sheet_(workbook: openpyxl.workbook.Workbook, property: ia.iaconfiguration._Property,
                              sheetpassword=None):
    """
    _append_validation_sheet(workbook, property, sheetpassword=None) -> Worksheet
    Appends to workbook a sheet containing data to be used in validation
    :param workbook: The workbook to which to append a sheet
    :param property: The iaconfiguration._Property object from which valid values are to be drawn
    :param sheetpassword: Optional password to be used in locking the sheet from editing
    :return: openpyxl.workbook.workbook.Worksheet
    """

    sheet_name = f"{property.display_name}_{property.property_id}"
    pwd = sheetpassword if sheetpassword is not None else sheet_name
    if sheet_name not in workbook.sheetnames:

        # First thing is to append a new sheet to the workbook to house the validation list
        # also, need to lock it so that the user can't change the values
        # not totally secure, but good enough for our purposes
        new_sheet = workbook.create_sheet(sheet_name)
        new_sheet.append(["property_value", "property_description", "property_value_id", "property_id"])

        # All data is stored as strings, but can be interpreted as either strings or numeric, depending on the
        # data_type attribute. Data needs to be stored in the appropriate format for validation purposes
        data_type = int if property.data_type == 'numeric' and property.decimal_digits == 0 else \
                    float if property.data_type == 'numeric' else str

        # Once the sheet has been appended, add the PropertyValue.value and PropertyValue.description to the sheet
        # for reference and validation
        for pv in property:  # type: ia.iaconfiguration._PropertyValue
            new_sheet.append((data_type(pv.value), pv.description, pv.property_value_id, pv.property_id))

        # Don't want users to mess with the validation values, so let's lock the worksheet
        # not really secure, as can be edited outside of excel, but for this purpose, it's good enough
        new_sheet.protection.enable()
        new_sheet.protection.password = pwd

        new_sheet.sheet_state = 'hidden'
        return new_sheet
    else:
        print(f"Sheet {sheet_name} already exists. Aborting")


def _set_global_validation_(data_sheet: openpyxl.workbook.workbook.Worksheet, row_number: int, vsheet):

    validation_title = vsheet.title
    dv = DataValidation(type="list",
                        formula1=f"{quote_sheetname(validation_title)}!$A$2:$A${vsheet.max_row}",
                        allow_blank=True
                        )
    dv.add(f"B{row_number}")
    data_sheet.add_data_validation(dv)


def _set_validation_(data_sheet: openpyxl.workbook.workbook.Worksheet, idx: str, vsheet, is_global=False):
    """
    Sets the validation for data_sheet, limiting values to those found in column A of vsheet
    :param data_sheet: The data sheet to which validation is to be applied
    :param idx: the column letter (if is_global=False) or row number (if is_global=True) of the cells to validate
    :param vsheet: The worksheet from which valid values are to be drawn (Drawn from column A)
    :param is_global: Whether data_sheet contains global data. Default False
    :return:
    """

    validation_title = vsheet.title

    # This DataValidation object specifies that a list rule is to be used,
    # drawing its values from column A of vsheet, and allowing null values
    dv = DataValidation(type="list",
                        formula1=f"{quote_sheetname(validation_title)}!$A$2:$A${vsheet.max_row}",
                        allow_blank=True
                       )

    # Cell address needs to be different for globals vs utterances
    # For globals, each row's B cell has it's own validation rule
    # For utterances, each column that has codes has its own validation rule.
    # Set cell address accordingly.
    column = idx if not is_global else "B"
    cell_address = f"B{idx}" if is_global else f"{column}2:{column}{data_sheet.max_row}"

    dv.add(cell_address)
    data_sheet.add_data_validation(dv)


def _build_global_sheet_(wb: openpyxl.Workbook, ia_config: ia.iaconfiguration._IaConfiguration):
    """
    _build_global_sheet(wb, ia_fonfig) -> Worksheet
    Constructs the worksheet into which global ratings will be entered
    :param wb: The workbook to which to append the worksheet
    :param ia_config: the caastools.parsing.ia.iaconfiguartion._IaConfiguration object specifying the globals to be scored
    :return: openpyxl.workbook.workbook.Worksheet
    """

    sheet = wb.create_sheet(title="Global Ratings", index=1)
    sheet.append(("Global", "Rating"))
    min_width = 18.0
    sheet.protection.enable()
    sheet.protection.password = "Globals"

    # insert a row for each global that needs to be scored
    for gp in ia_config.global_properties:  # type: ia.iaconfiguration._GlobalProperty
        sheet.append((gp.display_name,))

    # apply some styling to the spreadhseet
    col = sheet.column_dimensions[get_column_letter(1)]
    old_align = Alignment(horizontal=col.alignment.horizontal, vertical=col.alignment.vertical)
    col.alignment = Alignment(horizontal="general", vertical="bottom", shrinkToFit=True)
    if col.width < min_width:
        col.width = min_width
    col.alignment = old_align

    # Apply some formatting, and unlock data entry cells for modification
    for i, row in enumerate(sheet.rows, 1):
        for cell in row:
            cell.border = BORDER
        sheet[f"B{i}"].protection = Protection(locked=False, hidden=False)

    return sheet


def _build_interview_sheet_(wb: openpyxl.Workbook, ia_config: ia.iaconfiguration._IaConfiguration,
                            interview: ia.data._NewDataSet):

    text_column_number = 3 + len(ia_config.coding_properties) + 1

    sheet = wb.active
    sheet.title = "Interview"
    column_headers = ['Line', 'Utt', 'Role'] + [cp.display_name for cp in ia_config.coding_properties] + ['Text']
    sheet.append(column_headers)
    sheet.protection.enable()
    sheet.protection.password = "Interview"

    for i, utt in enumerate(interview.utterances, 1):  # type: ia.data._Utterances

        new_line = [utt.line_number, utt.utterance_number, utt.speaker_role] + \
                   [None for cp in ia_config.coding_properties] + [utt.utterance_text]
        sheet.append(new_line)

        # The entire sheet was locked above. Need to unprotect ONLY the data entry cells in each row
        # so that utterances can be coded
        for col, cp in enumerate(ia_config.coding_properties, 68):
            sheet[f"{chr(col)}{i}"].protection = Protection(locked=False, hidden=False)

    # Do the styling for the worksheet
    sheet.column_dimensions[get_column_letter(text_column_number)].width = 90.71
    sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
    sheet.page_setup.fitToHeight = 1
    sheet.page_setup.fitToWidth = 1

    for row in sheet.rows:
        i = 1
        for cell in row:
            if i % text_column_number == 0:
                cell.alignment = ALIGN
            cell.border = BORDER
            i += 1

    return sheet


def ia2xl(config_path, interview_files):
    """
    ia2xl(config_path, *interview_files) -> openpyxl.Workbook
    Converts an IA XML interview export to an excel workbook
    :param config_path: path to the IA coding system configuration that was used to score the interview
    :param interview_files: paths to files that contain the interview XML data
    :return: openpyxl.Workbook
    """
    property_column = 67
    global_row = 1

    # First, use caastools API to parse the IA configuration and the interview
    iacfg = ia.IaConfiguration(config_path)  # type: ia.iaconfiguration._IaConfiguration
    interview = ia.data.InterviewData("NewInterview", interview_files)

    wb = openpyxl.Workbook()

    # The first thing the workbook needs is a sheet to hold the interview
    # and another one to hold the global ratings
    utterances = _build_interview_sheet_(wb, iacfg, interview)
    global_data = _build_global_sheet_(wb, iacfg)


    """
    IA configurations, when parsed separate out into
    PropertyValue objects and GlobalValue objects
    Globals will need to be in their own worksheet
    but for each PropertyValue, we will need:
    1 - A worksheet in the workbook that defines the Data Validation characteristics of a column
    2 - A column in the Worksheet, named properly, which will be validated by the aforementioned sheet
    
    Globals will need:
    Their own spreadsheet into which data will be entered
    One worksheet with validation data for each GlobalProperty
    """

    for cp in iacfg.coding_properties + iacfg.global_properties:
        sheet = _append_validation_sheet_(wb, cp)

        is_global = isinstance(cp, ia.iaconfiguration._GlobalProperty)
        if is_global:
            global_row += 1
            idx = str(global_row)
            data_sheet = global_data
        else:
            property_column += 1
            idx = chr(property_column)
            data_sheet = utterances

        _set_validation_(data_sheet, idx, sheet, is_global)

    return wb
